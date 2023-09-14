# this is a python program for a 3x3x3 rubik's cube represented in excel
# importing user-defined turnAlgo.py library, which contains algorithms for turning
# Note: buy a rubik's cube


import openpyxl as pyxl
from openpyxl.styles import PatternFill 
import turnAlgo as turn


#loading an existing workbook, and assigning new sheet
workbook = pyxl.load_workbook("cube.xlsx")
sheet = workbook["Sheet1"]


#labelling the faces 'name (index)'
sheet['g1'].value  = 'TOP (4)'
sheet['c9'].value  = 'LEFT (0)'
sheet['g9'].value  = 'FRONT (1)'
sheet['k9'].value  = 'RIGHT (2)'
sheet['o9'].value  = 'BACK (3)'
sheet['g17'].value  = 'BOTTOM (5)'


#assign cells to faces
top_face = [sheet['f3'], sheet['g3'], sheet['h3'], 
            sheet['f4'], sheet['g4'], sheet['h4'], 
            sheet['f5'], sheet['g5'], sheet['h5']]     

left_face = [sheet['b11'], sheet['c11'], sheet['d11'], 
            sheet['b12'], sheet['c12'], sheet['d12'], 
            sheet['b13'], sheet['c13'], sheet['d13']]  

front_face = [sheet['f11'], sheet['g11'], sheet['h11'], 
            sheet['f12'], sheet['g12'], sheet['h12'], 
            sheet['f13'], sheet['g13'], sheet['h13']] 

right_face = [sheet['j11'], sheet['k11'], sheet['l11'], 
            sheet['j12'], sheet['k12'], sheet['l12'], 
            sheet['j13'], sheet['k13'], sheet['l13']]
            
back_face = [sheet['n11'], sheet['o11'], sheet['p11'], 
            sheet['n12'], sheet['o12'], sheet['p12'], 
            sheet['n13'], sheet['o13'], sheet['p13']]                

bottom_face = [sheet['f19'], sheet['g19'], sheet['h19'], 
            sheet['f20'], sheet['g20'], sheet['h20'], 
            sheet['f21'], sheet['g21'], sheet['h21']] 


#labelling colors hex code
yellow = "FFFF00"
white = "FFFFFF"
red = "FF0000"
blue = "0000FF"
orange = "FFA500"
green = "008000"


#function to color faces
def init_face(face, color):
    for cell in face:
        cell.fill = pyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.value = color
        

#color faces
init_face(front_face, white)  
init_face(top_face, red)                  
init_face(back_face, yellow)  
init_face(bottom_face, blue)  
init_face(left_face, orange)  
init_face(right_face, green) 


#create new cube
new_cube = (left_face, front_face, right_face, back_face, top_face, bottom_face, )


#play with cube by calling turn functions
turn.turn_counter(new_cube)
#turn.turn_right(new_cube)


#save changes in a new workbook
workbook.save('newCube.xlsx')