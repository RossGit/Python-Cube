#Algorithms for turns in a rubik's 
#Given an cube whith the dimensions n*n*n where n=3
#we can derive 18 moves from (3*2*n)
# where 3 is the number of axes(x,y,z) and 2 is the number of directions each axis can turn


import openpyxl as pyxl
from openpyxl.styles import PatternFill


def right_up(cube):
    temp1 = cube[1][2].value
    temp2 = cube[1][5].value
    temp3 = cube[1][8].value
    cube[1][2].fill = pyxl.styles.PatternFill(start_color=cube[5][2].value, end_color=cube[5][2].value, fill_type='solid')
    cube[1][2].value = cube[5][2].value
    cube[1][5].fill = pyxl.styles.PatternFill(start_color=cube[5][5].value, end_color=cube[5][5].value, fill_type='solid')
    cube[1][5].value = cube[5][5].value
    cube[1][8].fill = pyxl.styles.PatternFill(start_color=cube[5][8].value, end_color=cube[5][8].value, fill_type='solid')
    cube[1][8].value = cube[5][8].value
    cube[5][2].fill = pyxl.styles.PatternFill(start_color=cube[3][6].value, end_color=cube[3][6].value, fill_type='solid')
    cube[5][2].value = cube[3][6].value
    cube[5][5].fill = pyxl.styles.PatternFill(start_color=cube[3][3].value, end_color=cube[3][3].value, fill_type='solid')
    cube[5][5].value = cube[3][3].value
    cube[5][8].fill = pyxl.styles.PatternFill(start_color=cube[3][0].value, end_color=cube[3][0].value, fill_type='solid')
    cube[5][8].value = cube[3][0].value
    cube[3][0].fill = pyxl.styles.PatternFill(start_color=cube[4][8].value, end_color=cube[4][8].value, fill_type='solid')
    cube[3][0].value = cube[4][8].value
    cube[3][3].fill = pyxl.styles.PatternFill(start_color=cube[4][5].value, end_color=cube[4][5].value, fill_type='solid')
    cube[3][3].value = cube[4][5].value
    cube[3][6].fill = pyxl.styles.PatternFill(start_color=cube[4][2].value, end_color=cube[4][2].value, fill_type='solid')
    cube[3][6].value = cube[4][2].value
    cube[4][2].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[4][2].value = temp1
    cube[4][5].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[4][5].value = temp2
    cube[4][8].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[4][8].value = temp3

    #rigth face correction
    temp = cube[2][0].value
    cube[2][0].fill = pyxl.styles.PatternFill(start_color=cube[2][6].value, end_color=cube[2][6].value, fill_type='solid')
    cube[2][0].value = cube[2][6].value
    cube[2][6].fill = pyxl.styles.PatternFill(start_color=cube[2][8].value, end_color=cube[2][8].value, fill_type='solid')
    cube[2][6].value = cube[2][8].value
    cube[2][8].fill = pyxl.styles.PatternFill(start_color=cube[2][2].value, end_color=cube[2][2].value, fill_type='solid')
    cube[2][8].value = cube[2][2].value
    cube[2][2].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[2][2].value = temp
    temp  = cube[2][1].value
    cube[2][1].fill = pyxl.styles.PatternFill(start_color=cube[2][3].value, end_color=cube[2][3].value, fill_type='solid')
    cube[2][1].value = cube[2][3].value
    cube[2][3].fill = pyxl.styles.PatternFill(start_color=cube[2][7].value, end_color=cube[2][7].value, fill_type='solid')
    cube[2][3].value = cube[2][7].value
    cube[2][7].fill = pyxl.styles.PatternFill(start_color=cube[2][5].value, end_color=cube[2][5].value, fill_type='solid')
    cube[2][7].value = cube[2][5].value
    cube[2][5].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[2][5].value = temp

def right_down(cube):
    temp1 = cube[1][2].value
    temp2 = cube[1][5].value
    temp3 = cube[1][8].value
    cube[1][2].fill = pyxl.styles.PatternFill(start_color=cube[4][2].value, end_color=cube[4][2].value, fill_type='solid')
    cube[1][2].value = cube[4][2].value
    cube[1][5].fill = pyxl.styles.PatternFill(start_color=cube[4][5].value, end_color=cube[4][5].value, fill_type='solid')
    cube[1][5].value = cube[4][5].value
    cube[1][8].fill = pyxl.styles.PatternFill(start_color=cube[4][8].value, end_color=cube[4][8].value, fill_type='solid')
    cube[1][8].value = cube[4][8].value
    cube[4][2].fill = pyxl.styles.PatternFill(start_color=cube[3][6].value, end_color=cube[3][6].value, fill_type='solid')
    cube[4][2].value = cube[3][6].value
    cube[4][5].fill = pyxl.styles.PatternFill(start_color=cube[3][3].value, end_color=cube[3][3].value, fill_type='solid')
    cube[4][5].value = cube[3][3].value
    cube[4][8].fill = pyxl.styles.PatternFill(start_color=cube[3][0].value, end_color=cube[3][0].value, fill_type='solid')
    cube[4][8].value = cube[3][0].value
    cube[3][0].fill = pyxl.styles.PatternFill(start_color=cube[5][8].value, end_color=cube[5][8].value, fill_type='solid')
    cube[3][0].value = cube[5][8].value
    cube[3][3].fill = pyxl.styles.PatternFill(start_color=cube[5][5].value, end_color=cube[5][5].value, fill_type='solid')
    cube[3][3].value = cube[5][5].value
    cube[3][6].fill = pyxl.styles.PatternFill(start_color=cube[5][2].value, end_color=cube[5][2].value, fill_type='solid')
    cube[3][6].value = cube[5][2].value
    cube[5][2].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[5][2].value = temp1
    cube[5][5].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[5][5].value = temp2
    cube[5][8].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[5][8].value = temp3

    #right face correction
    temp = cube[2][0].value
    cube[2][0].fill = pyxl.styles.PatternFill(start_color=cube[2][2].value, end_color=cube[2][2].value, fill_type='solid')
    cube[2][0].value = cube[2][2].value
    cube[2][2].fill = pyxl.styles.PatternFill(start_color=cube[2][8].value, end_color=cube[2][8].value, fill_type='solid')
    cube[2][2].value = cube[2][8].value
    cube[2][8].fill = pyxl.styles.PatternFill(start_color=cube[2][6].value, end_color=cube[2][6].value, fill_type='solid')
    cube[2][8].value = cube[2][6].value
    cube[2][6].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[2][6].value = temp
    temp  = cube[2][1].value
    cube[2][1].fill = pyxl.styles.PatternFill(start_color=cube[2][5].value, end_color=cube[2][5].value, fill_type='solid')
    cube[2][1].value = cube[2][5].value
    cube[2][5].fill = pyxl.styles.PatternFill(start_color=cube[2][7].value, end_color=cube[2][7].value, fill_type='solid')
    cube[2][5].value = cube[2][7].value
    cube[2][7].fill = pyxl.styles.PatternFill(start_color=cube[2][3].value, end_color=cube[2][3].value, fill_type='solid')
    cube[2][7].value = cube[2][3].value
    cube[2][3].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[2][3].value = temp

def left_up(cube):
    temp1 = cube[1][0].value
    temp2 = cube[1][3].value
    temp3 = cube[1][6].value
    cube[1][0].fill = pyxl.styles.PatternFill(start_color=cube[5][0].value, end_color=cube[5][0].value, fill_type='solid')
    cube[1][0].value = cube[5][0].value
    cube[1][3].fill = pyxl.styles.PatternFill(start_color=cube[5][3].value, end_color=cube[5][3].value, fill_type='solid')
    cube[1][3].value = cube[5][3].value
    cube[1][6].fill = pyxl.styles.PatternFill(start_color=cube[5][6].value, end_color=cube[5][6].value, fill_type='solid')
    cube[1][6].value = cube[5][6].value
    cube[5][0].fill = pyxl.styles.PatternFill(start_color=cube[3][8].value, end_color=cube[3][8].value, fill_type='solid')
    cube[5][0].value = cube[3][8].value
    cube[5][3].fill = pyxl.styles.PatternFill(start_color=cube[3][5].value, end_color=cube[3][5].value, fill_type='solid')
    cube[5][3].value = cube[3][5].value
    cube[5][6].fill = pyxl.styles.PatternFill(start_color=cube[3][2].value, end_color=cube[3][2].value, fill_type='solid')
    cube[5][6].value = cube[3][2].value
    cube[3][8].fill = pyxl.styles.PatternFill(start_color=cube[4][0].value, end_color=cube[4][0].value, fill_type='solid')
    cube[3][8].value = cube[4][0].value
    cube[3][5].fill = pyxl.styles.PatternFill(start_color=cube[4][3].value, end_color=cube[4][3].value, fill_type='solid')
    cube[3][5].value = cube[4][3].value
    cube[3][2].fill = pyxl.styles.PatternFill(start_color=cube[4][6].value, end_color=cube[4][6].value, fill_type='solid')
    cube[3][2].value = cube[4][6].value
    cube[4][0].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[4][0].value = temp1
    cube[4][3].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[4][3].value = temp2
    cube[4][6].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[4][6].value = temp3

    #left face correction
    temp = cube[0][0].value
    cube[0][0].fill = pyxl.styles.PatternFill(start_color=cube[0][2].value, end_color=cube[0][2].value, fill_type='solid')
    cube[0][0].value = cube[0][2].value
    cube[0][2].fill = pyxl.styles.PatternFill(start_color=cube[0][8].value, end_color=cube[0][8].value, fill_type='solid')
    cube[0][2].value = cube[0][8].value
    cube[0][8].fill = pyxl.styles.PatternFill(start_color=cube[0][6].value, end_color=cube[0][6].value, fill_type='solid')
    cube[0][8].value = cube[0][6].value
    cube[0][6].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[0][6].value = temp
    temp  = cube[0][1].value
    cube[0][1].fill = pyxl.styles.PatternFill(start_color=cube[0][5].value, end_color=cube[0][5].value, fill_type='solid')
    cube[0][1].value = cube[0][5].value

    cube[0][5].fill = pyxl.styles.PatternFill(start_color=cube[0][7].value, end_color=cube[0][7].value, fill_type='solid')
    cube[0][5].value = cube[0][7].value
    cube[0][7].fill = pyxl.styles.PatternFill(start_color=cube[0][3].value, end_color=cube[0][3].value, fill_type='solid')
    cube[0][7].value = cube[0][3].value
    cube[0][3].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[0][3].value = temp

def left_down(cube):
    temp1 = cube[1][0].value
    temp2 = cube[1][3].value
    temp3 = cube[1][6].value
    cube[1][0].fill = pyxl.styles.PatternFill(start_color=cube[4][0].value, end_color=cube[4][0].value, fill_type='solid')
    cube[1][0].value = cube[4][0].value
    cube[1][3].fill = pyxl.styles.PatternFill(start_color=cube[4][3].value, end_color=cube[4][3].value, fill_type='solid')
    cube[1][3].value = cube[4][3].value
    cube[1][6].fill = pyxl.styles.PatternFill(start_color=cube[4][6].value, end_color=cube[4][6].value, fill_type='solid')
    cube[1][6].value = cube[4][6].value
    cube[4][0].fill = pyxl.styles.PatternFill(start_color=cube[3][8].value, end_color=cube[3][8].value, fill_type='solid')
    cube[4][0].value = cube[3][8].value
    cube[4][3].fill = pyxl.styles.PatternFill(start_color=cube[3][5].value, end_color=cube[3][5].value, fill_type='solid')
    cube[4][3].value = cube[3][5].value
    cube[4][6].fill = pyxl.styles.PatternFill(start_color=cube[3][2].value, end_color=cube[3][2].value, fill_type='solid')
    cube[4][6].value = cube[3][2].value
    cube[3][8].fill = pyxl.styles.PatternFill(start_color=cube[5][0].value, end_color=cube[5][0].value, fill_type='solid')
    cube[3][8].value = cube[5][0].value
    cube[3][5].fill = pyxl.styles.PatternFill(start_color=cube[5][3].value, end_color=cube[5][3].value, fill_type='solid')
    cube[3][5].value = cube[5][3].value
    cube[3][2].fill = pyxl.styles.PatternFill(start_color=cube[5][6].value, end_color=cube[5][6].value, fill_type='solid')
    cube[3][2].value = cube[5][6].value
    cube[5][0].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[5][0].value = temp1
    cube[5][3].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[5][3].value = temp2
    cube[5][6].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[5][6].value = temp3

    #left face correction#
    temp = cube[0][0].value
    cube[0][0].fill = pyxl.styles.PatternFill(start_color=cube[0][6].value, end_color=cube[0][6].value, fill_type='solid')
    cube[0][0].value = cube[0][6].value
    cube[0][6].fill = pyxl.styles.PatternFill(start_color=cube[0][8].value, end_color=cube[0][8].value, fill_type='solid')
    cube[0][6].value = cube[0][8].value
    cube[0][8].fill = pyxl.styles.PatternFill(start_color=cube[0][2].value, end_color=cube[0][2].value, fill_type='solid')
    cube[0][8].value = cube[0][2].value
    cube[0][2].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[0][2].value = temp
    temp  = cube[0][1].value
    cube[0][1].fill = pyxl.styles.PatternFill(start_color=cube[0][3].value, end_color=cube[0][3].value, fill_type='solid')
    cube[0][1].value = cube[0][3].value
    cube[0][3].fill = pyxl.styles.PatternFill(start_color=cube[0][7].value, end_color=cube[0][7].value, fill_type='solid')
    cube[0][3].value = cube[0][7].value
    cube[0][7].fill = pyxl.styles.PatternFill(start_color=cube[0][5].value, end_color=cube[0][5].value, fill_type='solid')
    cube[0][7].value = cube[0][5].value
    cube[0][5].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[0][5].value = temp

def top_right(cube):
    temp1 = cube[1][0].value
    temp2 = cube[1][1].value
    temp3 = cube[1][2].value
    cube[1][0].fill = pyxl.styles.PatternFill(start_color=cube[0][0].value, end_color=cube[0][0].value, fill_type='solid')
    cube[1][0].value = cube[0][0].value
    cube[1][1].fill = pyxl.styles.PatternFill(start_color=cube[0][1].value, end_color=cube[0][1].value, fill_type='solid')
    cube[1][1].value = cube[0][1].value
    cube[1][2].fill = pyxl.styles.PatternFill(start_color=cube[0][2].value, end_color=cube[0][2].value, fill_type='solid')
    cube[1][2].value = cube[0][2].value
    cube[0][0].fill = pyxl.styles.PatternFill(start_color=cube[3][0].value, end_color=cube[3][0].value, fill_type='solid')
    cube[0][0].value = cube[3][0].value
    cube[0][1].fill = pyxl.styles.PatternFill(start_color=cube[3][1].value, end_color=cube[3][1].value, fill_type='solid')
    cube[0][1].value = cube[3][1].value
    cube[0][2].fill = pyxl.styles.PatternFill(start_color=cube[3][2].value, end_color=cube[3][2].value, fill_type='solid')
    cube[0][2].value = cube[3][2].value
    cube[3][0].fill = pyxl.styles.PatternFill(start_color=cube[2][0].value, end_color=cube[2][0].value, fill_type='solid')
    cube[3][0].value = cube[2][0].value
    cube[3][1].fill = pyxl.styles.PatternFill(start_color=cube[2][1].value, end_color=cube[2][1].value, fill_type='solid')
    cube[3][1].value = cube[2][1].value
    cube[3][2].fill = pyxl.styles.PatternFill(start_color=cube[2][2].value, end_color=cube[2][2].value, fill_type='solid')
    cube[3][2].value = cube[2][2].value
    cube[2][0].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[2][0].value = temp1
    cube[2][1].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[2][1].value = temp2
    cube[2][2].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[2][2].value = temp3

    #top face correction#
    temp = cube[4][0].value
    cube[4][0].fill = pyxl.styles.PatternFill(start_color=cube[4][2].value, end_color=cube[4][2].value, fill_type='solid')
    cube[4][0].value = cube[4][2].value
    cube[4][2].fill = pyxl.styles.PatternFill(start_color=cube[4][8].value, end_color=cube[4][8].value, fill_type='solid')
    cube[4][2].value = cube[4][8].value
    cube[4][8].fill = pyxl.styles.PatternFill(start_color=cube[4][6].value, end_color=cube[4][6].value, fill_type='solid')
    cube[4][8].value = cube[4][6].value
    cube[4][6].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[4][6].value = temp
    temp = cube[4][1].value
    cube[4][1].fill = pyxl.styles.PatternFill(start_color=cube[4][5].value, end_color=cube[4][5].value, fill_type='solid')
    cube[4][1].value = cube[4][5].value
    cube[4][5].fill = pyxl.styles.PatternFill(start_color=cube[4][7].value, end_color=cube[4][7].value, fill_type='solid')
    cube[4][5].value = cube[4][7].value
    cube[4][7].fill = pyxl.styles.PatternFill(start_color=cube[4][3].value, end_color=cube[4][3].value, fill_type='solid')
    cube[4][7].value = cube[4][3].value
    cube[4][3].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[4][3].value = temp

def top_left(cube):
    temp1 = cube[1][0].value
    temp2 = cube[1][1].value
    temp3 = cube[1][2].value
    cube[1][0].fill = pyxl.styles.PatternFill(start_color=cube[2][0].value, end_color=cube[2][0].value, fill_type='solid')
    cube[1][0].value = cube[2][0].value
    cube[1][1].fill = pyxl.styles.PatternFill(start_color=cube[2][1].value, end_color=cube[2][1].value, fill_type='solid')
    cube[1][1].value = cube[2][1].value
    cube[1][2].fill = pyxl.styles.PatternFill(start_color=cube[2][2].value, end_color=cube[2][2].value, fill_type='solid')
    cube[1][2].value = cube[2][2].value
    cube[2][0].fill = pyxl.styles.PatternFill(start_color=cube[3][0].value, end_color=cube[3][0].value, fill_type='solid')
    cube[2][0].value = cube[3][0].value
    cube[2][1].fill = pyxl.styles.PatternFill(start_color=cube[3][1].value, end_color=cube[3][1].value, fill_type='solid')
    cube[2][1].value = cube[3][1].value
    cube[2][2].fill = pyxl.styles.PatternFill(start_color=cube[3][2].value, end_color=cube[3][2].value, fill_type='solid')
    cube[2][2].value = cube[3][2].value
    cube[3][0].fill = pyxl.styles.PatternFill(start_color=cube[0][0].value, end_color=cube[0][0].value, fill_type='solid')
    cube[3][0].value = cube[0][0].value
    cube[3][1].fill = pyxl.styles.PatternFill(start_color=cube[0][1].value, end_color=cube[0][1].value, fill_type='solid')
    cube[3][1].value = cube[0][1].value
    cube[3][2].fill = pyxl.styles.PatternFill(start_color=cube[0][2].value, end_color=cube[0][2].value, fill_type='solid')
    cube[3][2].value = cube[0][2].value
    cube[0][0].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[0][0].value = temp1
    cube[0][1].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[0][1].value = temp2
    cube[0][2].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[0][2].value = temp3
    #top face correction#
    temp = cube[4][0].value
    cube[4][0].fill = pyxl.styles.PatternFill(start_color=cube[4][6].value, end_color=cube[4][6].value, fill_type='solid')
    cube[4][0].value = cube[4][6].value
    cube[4][6].fill = pyxl.styles.PatternFill(start_color=cube[4][8].value, end_color=cube[4][8].value, fill_type='solid')
    cube[4][6].value = cube[4][8].value
    cube[4][8].fill = pyxl.styles.PatternFill(start_color=cube[4][2].value, end_color=cube[4][2].value, fill_type='solid')
    cube[4][8].value = cube[4][2].value
    cube[4][2].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[4][2].value = temp
    temp = cube[4][1].value
    cube[4][1].fill = pyxl.styles.PatternFill(start_color=cube[4][3].value, end_color=cube[4][3].value, fill_type='solid')
    cube[4][1].value = cube[4][3].value
    cube[4][3].fill = pyxl.styles.PatternFill(start_color=cube[4][7].value, end_color=cube[4][7].value, fill_type='solid')
    cube[4][3].value = cube[4][7].value
    cube[4][7].fill = pyxl.styles.PatternFill(start_color=cube[4][5].value, end_color=cube[4][5].value, fill_type='solid')
    cube[4][7].value = cube[4][5].value
    cube[4][5].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[4][5].value = temp

def bottom_right(cube):
    temp1 = cube[1][6].value
    temp2 = cube[1][7].value
    temp3 = cube[1][8].value
    cube[1][6].fill = pyxl.styles.PatternFill(start_color=cube[0][6].value, end_color=cube[0][6].value, fill_type='solid')
    cube[1][6].value = cube[0][6].value
    cube[1][7].fill = pyxl.styles.PatternFill(start_color=cube[0][7].value, end_color=cube[0][7].value, fill_type='solid')
    cube[1][7].value = cube[0][7].value
    cube[1][8].fill = pyxl.styles.PatternFill(start_color=cube[0][8].value, end_color=cube[0][8].value, fill_type='solid')
    cube[1][8].value = cube[0][8].value
    cube[0][6].fill = pyxl.styles.PatternFill(start_color=cube[3][6].value, end_color=cube[3][6].value, fill_type='solid')
    cube[0][6].value = cube[3][6].value
    cube[0][7].fill = pyxl.styles.PatternFill(start_color=cube[3][7].value, end_color=cube[3][7].value, fill_type='solid')
    cube[0][7].value = cube[3][7].value
    cube[0][8].fill = pyxl.styles.PatternFill(start_color=cube[3][8].value, end_color=cube[3][8].value, fill_type='solid')
    cube[0][8].value = cube[3][8].value
    cube[3][6].fill = pyxl.styles.PatternFill(start_color=cube[2][6].value, end_color=cube[2][6].value, fill_type='solid')
    cube[3][6].value = cube[2][6].value
    cube[3][7].fill = pyxl.styles.PatternFill(start_color=cube[2][7].value, end_color=cube[2][7].value, fill_type='solid')
    cube[3][7].value = cube[2][7].value
    cube[3][8].fill = pyxl.styles.PatternFill(start_color=cube[2][8].value, end_color=cube[2][8].value, fill_type='solid')
    cube[3][8].value = cube[2][8].value
    cube[2][6].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[2][6].value = temp1
    cube[2][7].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[2][7].value = temp2
    cube[2][8].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[2][8].value = temp3
    #bottom face correction
    temp = cube[5][0].value
    cube[5][0].fill = pyxl.styles.PatternFill(start_color=cube[5][6].value, end_color=cube[5][6].value, fill_type='solid')
    cube[5][0].value = cube[5][6].value
    cube[5][6].fill = pyxl.styles.PatternFill(start_color=cube[5][8].value, end_color=cube[5][8].value, fill_type='solid')
    cube[5][6].value = cube[5][8].value
    cube[5][8].fill = pyxl.styles.PatternFill(start_color=cube[5][2].value, end_color=cube[5][2].value, fill_type='solid')
    cube[5][8].value = cube[5][2].value
    cube[5][2].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[5][2].value = temp
    temp = cube[5][1].value
    cube[5][1].fill = pyxl.styles.PatternFill(start_color=cube[5][3].value, end_color=cube[5][3].value, fill_type='solid')
    cube[5][1].value = cube[5][3].value
    cube[5][3].fill = pyxl.styles.PatternFill(start_color=cube[5][7].value, end_color=cube[5][7].value, fill_type='solid')
    cube[5][3].value = cube[5][7].value
    cube[5][7].fill = pyxl.styles.PatternFill(start_color=cube[5][5].value, end_color=cube[5][5].value, fill_type='solid')
    cube[5][7].value = cube[5][5].value
    cube[5][5].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[5][5].value = temp

def bottom_left(cube):
    temp1 = cube[1][6].value
    temp2 = cube[1][7].value
    temp3 = cube[1][8].value
    cube[1][6].fill = pyxl.styles.PatternFill(start_color=cube[2][6].value, end_color=cube[2][6].value, fill_type='solid')
    cube[1][6].value = cube[2][6].value
    cube[1][7].fill = pyxl.styles.PatternFill(start_color=cube[2][7].value, end_color=cube[2][7].value, fill_type='solid')
    cube[1][7].value = cube[2][7].value
    cube[1][8].fill = pyxl.styles.PatternFill(start_color=cube[2][8].value, end_color=cube[2][8].value, fill_type='solid')
    cube[1][8].value = cube[2][8].value
    cube[2][6].fill = pyxl.styles.PatternFill(start_color=cube[3][6].value, end_color=cube[3][6].value, fill_type='solid')
    cube[2][6].value = cube[3][6].value
    cube[2][7].fill = pyxl.styles.PatternFill(start_color=cube[3][7].value, end_color=cube[3][7].value, fill_type='solid')
    cube[2][7].value = cube[3][7].value
    cube[2][8].fill = pyxl.styles.PatternFill(start_color=cube[3][8].value, end_color=cube[3][8].value, fill_type='solid')
    cube[2][8].value = cube[3][8].value
    cube[3][6].fill = pyxl.styles.PatternFill(start_color=cube[0][6].value, end_color=cube[0][6].value, fill_type='solid')
    cube[3][6].value = cube[0][6].value
    cube[3][7].fill = pyxl.styles.PatternFill(start_color=cube[0][7].value, end_color=cube[0][7].value, fill_type='solid')
    cube[3][7].value = cube[0][7].value
    cube[3][8].fill = pyxl.styles.PatternFill(start_color=cube[0][8].value, end_color=cube[0][8].value, fill_type='solid')
    cube[3][8].value = cube[0][8].value
    cube[0][6].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[0][6].value = temp1
    cube[0][7].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[0][7].value = temp2
    cube[0][8].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[0][8].value = temp3
    #bottom face correction
    temp = cube[5][0].value
    cube[5][0].fill = pyxl.styles.PatternFill(start_color=cube[5][2].value, end_color=cube[5][2].value, fill_type='solid')
    cube[5][0].value = cube[5][2].value
    cube[5][2].fill = pyxl.styles.PatternFill(start_color=cube[5][8].value, end_color=cube[5][8].value, fill_type='solid')
    cube[5][2].value = cube[5][8].value
    cube[5][8].fill = pyxl.styles.PatternFill(start_color=cube[5][6].value, end_color=cube[5][6].value, fill_type='solid')
    cube[5][8].value = cube[5][6].value
    cube[5][6].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[5][6].value = temp
    temp = cube[5][1].value
    cube[5][1].fill = pyxl.styles.PatternFill(start_color=cube[5][5].value, end_color=cube[5][5].value, fill_type='solid')
    cube[5][1].value = cube[5][5].value
    cube[5][5].fill = pyxl.styles.PatternFill(start_color=cube[5][7].value, end_color=cube[5][7].value, fill_type='solid')
    cube[5][5].value = cube[5][7].value
    cube[5][7].fill = pyxl.styles.PatternFill(start_color=cube[5][3].value, end_color=cube[5][3].value, fill_type='solid')
    cube[5][7].value = cube[5][3].value
    cube[5][3].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[5][3].value = temp

def center_up(cube):
    temp1 = cube[1][1].value
    temp2 = cube[1][4].value
    temp3 = cube[1][7].value
    cube[1][1].fill = pyxl.styles.PatternFill(start_color=cube[5][1].value, end_color=cube[5][1].value, fill_type='solid')
    cube[1][1].value = cube[5][1].value
    cube[1][4].fill = pyxl.styles.PatternFill(start_color=cube[5][4].value, end_color=cube[5][4].value, fill_type='solid')
    cube[1][4].value = cube[5][4].value
    cube[1][7].fill = pyxl.styles.PatternFill(start_color=cube[5][7].value, end_color=cube[5][7].value, fill_type='solid')
    cube[1][7].value = cube[5][7].value
    cube[5][1].fill = pyxl.styles.PatternFill(start_color=cube[3][7].value, end_color=cube[3][7].value, fill_type='solid')
    cube[5][1].value = cube[3][7].value
    cube[5][4].fill = pyxl.styles.PatternFill(start_color=cube[3][4].value, end_color=cube[3][4].value, fill_type='solid')
    cube[5][4].value = cube[3][4].value
    cube[5][7].fill = pyxl.styles.PatternFill(start_color=cube[3][1].value, end_color=cube[3][1].value, fill_type='solid')
    cube[5][7].value = cube[3][1].value
    cube[3][7].fill = pyxl.styles.PatternFill(start_color=cube[4][1].value, end_color=cube[4][1].value, fill_type='solid')
    cube[3][7].value = cube[4][1].value
    cube[3][4].fill = pyxl.styles.PatternFill(start_color=cube[4][4].value, end_color=cube[4][4].value, fill_type='solid')
    cube[3][4].value = cube[4][4].value
    cube[3][1].fill = pyxl.styles.PatternFill(start_color=cube[4][7].value, end_color=cube[4][7].value, fill_type='solid')
    cube[3][1].value = cube[4][7].value
    cube[4][1].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[4][1].value = temp1
    cube[4][4].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[4][4].value = temp2
    cube[4][7].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[4][7].value = temp3

def center_down(cube):
    temp1 = cube[1][1].value
    temp2 = cube[1][4].value
    temp3 = cube[1][7].value
    cube[1][1].fill = pyxl.styles.PatternFill(start_color=cube[4][1].value, end_color=cube[4][1].value, fill_type='solid')
    cube[1][1].value = cube[4][1].value
    cube[1][4].fill = pyxl.styles.PatternFill(start_color=cube[4][4].value, end_color=cube[4][4].value, fill_type='solid')
    cube[1][4].value = cube[4][4].value
    cube[1][7].fill = pyxl.styles.PatternFill(start_color=cube[4][7].value, end_color=cube[4][7].value, fill_type='solid')
    cube[1][7].value = cube[4][7].value
    cube[4][1].fill = pyxl.styles.PatternFill(start_color=cube[3][7].value, end_color=cube[3][7].value, fill_type='solid')
    cube[4][1].value = cube[3][7].value
    cube[4][4].fill = pyxl.styles.PatternFill(start_color=cube[3][4].value, end_color=cube[3][4].value, fill_type='solid')
    cube[4][4].value = cube[3][4].value
    cube[4][7].fill = pyxl.styles.PatternFill(start_color=cube[3][1].value, end_color=cube[3][1].value, fill_type='solid')
    cube[4][7].value = cube[3][1].value
    cube[3][7].fill = pyxl.styles.PatternFill(start_color=cube[5][1].value, end_color=cube[5][1].value, fill_type='solid')
    cube[3][7].value = cube[5][1].value
    cube[3][4].fill = pyxl.styles.PatternFill(start_color=cube[5][4].value, end_color=cube[5][4].value, fill_type='solid')
    cube[3][4].value = cube[5][4].value
    cube[3][1].fill = pyxl.styles.PatternFill(start_color=cube[5][7].value, end_color=cube[5][7].value, fill_type='solid')
    cube[3][1].value = cube[5][7].value
    cube[5][1].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[5][1].value = temp1
    cube[5][4].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[5][4].value = temp2
    cube[5][7].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[5][7].value = temp3

def middle_left(cube):
    temp1 = cube[1][3].value
    temp2 = cube[1][4].value
    temp3 = cube[1][5].value
    cube[1][3].fill = pyxl.styles.PatternFill(start_color=cube[2][3].value, end_color=cube[2][3].value, fill_type='solid')
    cube[1][3].value = cube[2][3].value
    cube[1][4].fill = pyxl.styles.PatternFill(start_color=cube[2][4].value, end_color=cube[2][4].value, fill_type='solid')
    cube[1][4].value = cube[2][4].value
    cube[1][5].fill = pyxl.styles.PatternFill(start_color=cube[2][5].value, end_color=cube[2][5].value, fill_type='solid')
    cube[1][5].value = cube[2][5].value
    cube[2][3].fill = pyxl.styles.PatternFill(start_color=cube[3][3].value, end_color=cube[3][3].value, fill_type='solid')
    cube[2][3].value = cube[3][3].value
    cube[2][4].fill = pyxl.styles.PatternFill(start_color=cube[3][4].value, end_color=cube[3][4].value, fill_type='solid')
    cube[2][4].value = cube[3][4].value
    cube[2][5].fill = pyxl.styles.PatternFill(start_color=cube[3][5].value, end_color=cube[3][5].value, fill_type='solid')
    cube[2][5].value = cube[3][5].value
    cube[3][3].fill = pyxl.styles.PatternFill(start_color=cube[0][3].value, end_color=cube[0][3].value, fill_type='solid')
    cube[3][3].value = cube[0][3].value
    cube[3][4].fill = pyxl.styles.PatternFill(start_color=cube[0][4].value, end_color=cube[0][4].value, fill_type='solid')
    cube[3][4].value = cube[0][4].value
    cube[3][5].fill = pyxl.styles.PatternFill(start_color=cube[0][5].value, end_color=cube[0][5].value, fill_type='solid')
    cube[3][5].value = cube[0][5].value
    cube[0][3].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[0][3].value = temp1
    cube[0][4].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[0][4].value = temp2
    cube[0][5].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[0][5].value = temp3

def middle_right(cube):
    temp1 = cube[1][3].value
    temp2 = cube[1][4].value
    temp3 = cube[1][5].value
    cube[1][3].fill = pyxl.styles.PatternFill(start_color=cube[0][3].value, end_color=cube[0][3].value, fill_type='solid')
    cube[1][3].value = cube[0][3].value
    cube[1][4].fill = pyxl.styles.PatternFill(start_color=cube[0][4].value, end_color=cube[0][4].value, fill_type='solid')
    cube[1][4].value = cube[0][4].value
    cube[1][5].fill = pyxl.styles.PatternFill(start_color=cube[0][5].value, end_color=cube[0][5].value, fill_type='solid')
    cube[1][5].value = cube[0][5].value
    cube[0][3].fill = pyxl.styles.PatternFill(start_color=cube[3][3].value, end_color=cube[3][3].value, fill_type='solid')
    cube[0][3].value = cube[3][3].value
    cube[0][4].fill = pyxl.styles.PatternFill(start_color=cube[3][4].value, end_color=cube[3][4].value, fill_type='solid')
    cube[0][4].value = cube[3][4].value
    cube[0][5].fill = pyxl.styles.PatternFill(start_color=cube[3][5].value, end_color=cube[3][5].value, fill_type='solid')
    cube[0][5].value = cube[3][5].value
    cube[3][3].fill = pyxl.styles.PatternFill(start_color=cube[2][3].value, end_color=cube[2][3].value, fill_type='solid')
    cube[3][3].value = cube[2][3].value
    cube[3][4].fill = pyxl.styles.PatternFill(start_color=cube[2][4].value, end_color=cube[2][4].value, fill_type='solid')
    cube[3][4].value = cube[2][4].value
    cube[3][5].fill = pyxl.styles.PatternFill(start_color=cube[2][5].value, end_color=cube[2][5].value, fill_type='solid')
    cube[3][5].value = cube[2][5].value
    cube[2][3].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[2][3].value = temp1
    cube[2][4].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[2][4].value = temp2
    cube[2][5].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[2][5].value = temp3

def core_clockwise(cube):
    temp1 = cube[4][3].value
    temp2 = cube[4][4].value
    temp3 = cube[4][5].value
    cube[4][3].fill = pyxl.styles.PatternFill(start_color=cube[0][7].value, end_color=cube[0][7].value, fill_type='solid')
    cube[4][3].value = cube[0][7].value
    cube[4][4].fill = pyxl.styles.PatternFill(start_color=cube[0][4].value, end_color=cube[0][4].value, fill_type='solid')
    cube[4][4].value = cube[0][4].value
    cube[4][5].fill = pyxl.styles.PatternFill(start_color=cube[0][1].value, end_color=cube[0][1].value, fill_type='solid')
    cube[4][5].value = cube[0][1].value
    cube[0][7].fill = pyxl.styles.PatternFill(start_color=cube[5][5].value, end_color=cube[5][5].value, fill_type='solid')
    cube[0][7].value = cube[5][5].value
    cube[0][4].fill = pyxl.styles.PatternFill(start_color=cube[5][4].value, end_color=cube[5][4].value, fill_type='solid')
    cube[0][4].value = cube[5][4].value
    cube[0][1].fill = pyxl.styles.PatternFill(start_color=cube[5][3].value, end_color=cube[5][3].value, fill_type='solid')
    cube[0][1].value = cube[5][3].value
    cube[5][5].fill = pyxl.styles.PatternFill(start_color=cube[2][1].value, end_color=cube[2][1].value, fill_type='solid')
    cube[5][5].value = cube[2][1].value
    cube[5][4].fill = pyxl.styles.PatternFill(start_color=cube[2][4].value, end_color=cube[2][4].value, fill_type='solid')
    cube[5][4].value = cube[2][4].value
    cube[5][3].fill = pyxl.styles.PatternFill(start_color=cube[2][7].value, end_color=cube[2][7].value, fill_type='solid')
    cube[5][3].value = cube[2][7].value
    cube[2][1].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[2][1].value = temp1
    cube[2][4].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[2][4].value = temp2
    cube[2][7].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[2][7].value = temp3

def core_counter(cube):
    temp1 = cube[4][3].value
    temp2 = cube[4][4].value
    temp3 = cube[4][5].value
    cube[4][3].fill = pyxl.styles.PatternFill(start_color=cube[2][1].value, end_color=cube[2][1].value, fill_type='solid')
    cube[4][3].value = cube[2][1].value
    cube[4][4].fill = pyxl.styles.PatternFill(start_color=cube[2][4].value, end_color=cube[2][4].value, fill_type='solid')
    cube[4][4].value = cube[2][4].value
    cube[4][5].fill = pyxl.styles.PatternFill(start_color=cube[2][7].value, end_color=cube[2][7].value, fill_type='solid')
    cube[4][5].value = cube[2][7].value
    cube[2][1].fill = pyxl.styles.PatternFill(start_color=cube[5][5].value, end_color=cube[5][5].value, fill_type='solid')
    cube[2][1].value = cube[5][5].value
    cube[2][4].fill = pyxl.styles.PatternFill(start_color=cube[5][4].value, end_color=cube[5][4].value, fill_type='solid')
    cube[2][4].value = cube[5][4].value
    cube[2][7].fill = pyxl.styles.PatternFill(start_color=cube[5][3].value, end_color=cube[5][3].value, fill_type='solid')
    cube[2][7].value = cube[5][3].value
    cube[5][5].fill = pyxl.styles.PatternFill(start_color=cube[0][7].value, end_color=cube[0][7].value, fill_type='solid')
    cube[5][5].value = cube[0][7].value
    cube[5][4].fill = pyxl.styles.PatternFill(start_color=cube[0][4].value, end_color=cube[0][4].value, fill_type='solid')
    cube[5][4].value = cube[0][4].value
    cube[5][3].fill = pyxl.styles.PatternFill(start_color=cube[0][1].value, end_color=cube[0][1].value, fill_type='solid')
    cube[5][3].value = cube[0][1].value
    cube[0][7].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[0][7].value = temp1
    cube[0][4].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[0][4].value = temp2
    cube[0][1].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[0][1].value = temp3

def front_clockwise(cube):
    temp1 = cube[4][6].value
    temp2 = cube[4][7].value
    temp3 = cube[4][8].value
    cube[4][6].fill = pyxl.styles.PatternFill(start_color=cube[0][8].value, end_color=cube[0][8].value, fill_type='solid')
    cube[4][6].value = cube[0][8].value
    cube[4][7].fill = pyxl.styles.PatternFill(start_color=cube[0][5].value, end_color=cube[0][5].value, fill_type='solid')
    cube[4][7].value = cube[0][5].value
    cube[4][8].fill = pyxl.styles.PatternFill(start_color=cube[0][2].value, end_color=cube[0][2].value, fill_type='solid')
    cube[4][8].value = cube[0][2].value
    cube[0][8].fill = pyxl.styles.PatternFill(start_color=cube[5][2].value, end_color=cube[5][2].value, fill_type='solid')
    cube[0][8].value = cube[5][2].value
    cube[0][5].fill = pyxl.styles.PatternFill(start_color=cube[5][1].value, end_color=cube[5][1].value, fill_type='solid')
    cube[0][5].value = cube[5][1].value
    cube[0][2].fill = pyxl.styles.PatternFill(start_color=cube[5][0].value, end_color=cube[5][0].value, fill_type='solid')
    cube[0][2].value = cube[5][0].value
    cube[5][2].fill = pyxl.styles.PatternFill(start_color=cube[2][0].value, end_color=cube[2][0].value, fill_type='solid')
    cube[5][2].value = cube[2][0].value
    cube[5][1].fill = pyxl.styles.PatternFill(start_color=cube[2][3].value, end_color=cube[2][3].value, fill_type='solid')
    cube[5][1].value = cube[2][3].value
    cube[5][0].fill = pyxl.styles.PatternFill(start_color=cube[2][6].value, end_color=cube[2][6].value, fill_type='solid')
    cube[5][0].value = cube[2][6].value
    cube[2][0].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[2][0].value = temp1
    cube[2][3].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[2][3].value = temp2
    cube[2][6].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[2][6].value = temp3

    #front face correction
    temp = cube[1][0].value
    cube[1][0].fill = pyxl.styles.PatternFill(start_color=cube[1][6].value, end_color=cube[1][6].value, fill_type='solid')
    cube[1][0].value = cube[1][6].value
    cube[1][6].fill = pyxl.styles.PatternFill(start_color=cube[1][8].value, end_color=cube[1][8].value, fill_type='solid')
    cube[1][6].value = cube[1][8].value
    cube[1][8].fill = pyxl.styles.PatternFill(start_color=cube[1][2].value, end_color=cube[1][2].value, fill_type='solid')
    cube[1][8].value = cube[1][2].value
    cube[1][2].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[1][2].value = temp
    temp = cube[1][1].value
    cube[1][1].fill = pyxl.styles.PatternFill(start_color=cube[1][3].value, end_color=cube[1][3].value, fill_type='solid')
    cube[1][1].value = cube[1][3].value
    cube[1][3].fill = pyxl.styles.PatternFill(start_color=cube[1][7].value, end_color=cube[1][7].value, fill_type='solid')
    cube[1][3].value = cube[1][7].value
    cube[1][7].fill = pyxl.styles.PatternFill(start_color=cube[1][5].value, end_color=cube[1][5].value, fill_type='solid')
    cube[1][7].value = cube[1][5].value
    cube[1][5].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[1][5].value = temp

def front_counter(cube):
    temp1 = cube[4][6].value
    temp2 = cube[4][7].value
    temp3 = cube[4][8].value
    cube[4][6].fill = pyxl.styles.PatternFill(start_color=cube[2][0].value, end_color=cube[2][0].value, fill_type='solid')
    cube[4][6].value = cube[2][0].value
    cube[4][7].fill = pyxl.styles.PatternFill(start_color=cube[2][3].value, end_color=cube[2][3].value, fill_type='solid')
    cube[4][7].value = cube[2][3].value
    cube[4][8].fill = pyxl.styles.PatternFill(start_color=cube[2][6].value, end_color=cube[2][6].value, fill_type='solid')
    cube[4][8].value = cube[2][6].value
    cube[2][0].fill = pyxl.styles.PatternFill(start_color=cube[5][2].value, end_color=cube[5][2].value, fill_type='solid')
    cube[2][0].value = cube[5][2].value
    cube[2][3].fill = pyxl.styles.PatternFill(start_color=cube[5][1].value, end_color=cube[5][1].value, fill_type='solid')
    cube[2][3].value = cube[5][1].value
    cube[2][6].fill = pyxl.styles.PatternFill(start_color=cube[5][0].value, end_color=cube[5][0].value, fill_type='solid')
    cube[2][6].value = cube[5][0].value
    cube[5][2].fill = pyxl.styles.PatternFill(start_color=cube[0][8].value, end_color=cube[0][8].value, fill_type='solid')
    cube[5][2].value = cube[0][8].value
    cube[5][1].fill = pyxl.styles.PatternFill(start_color=cube[0][5].value, end_color=cube[0][5].value, fill_type='solid')
    cube[5][1].value = cube[0][5].value
    cube[5][0].fill = pyxl.styles.PatternFill(start_color=cube[0][2].value, end_color=cube[0][2].value, fill_type='solid')
    cube[5][0].value = cube[0][2].value
    cube[0][8].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[0][8].value = temp1
    cube[0][5].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[0][5].value = temp2
    cube[0][2].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[0][2].value = temp3

    #front face correction
    temp = cube[1][0].value
    cube[1][0].fill = pyxl.styles.PatternFill(start_color=cube[1][2].value, end_color=cube[1][2].value, fill_type='solid')
    cube[1][0].value = cube[1][2].value
    cube[1][2].fill = pyxl.styles.PatternFill(start_color=cube[1][8].value, end_color=cube[1][8].value, fill_type='solid')
    cube[1][2].value = cube[1][8].value
    cube[1][8].fill = pyxl.styles.PatternFill(start_color=cube[1][6].value, end_color=cube[1][6].value, fill_type='solid')
    cube[1][8].value = cube[1][6].value
    cube[1][6].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[1][6].value = temp
    temp = cube[1][1].value
    cube[1][1].fill = pyxl.styles.PatternFill(start_color=cube[1][5].value, end_color=cube[1][5].value, fill_type='solid')
    cube[1][1].value = cube[1][5].value
    cube[1][5].fill = pyxl.styles.PatternFill(start_color=cube[1][7].value, end_color=cube[1][7].value, fill_type='solid')
    cube[1][5].value = cube[1][7].value
    cube[1][7].fill = pyxl.styles.PatternFill(start_color=cube[1][3].value, end_color=cube[1][3].value, fill_type='solid')
    cube[1][7].value = cube[1][3].value
    cube[1][3].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[1][3].value = temp

def back_clockwise(cube):
    temp1 = cube[4][0].value
    temp2 = cube[4][1].value
    temp3 = cube[4][2].value
    cube[4][0].fill = pyxl.styles.PatternFill(start_color=cube[0][6].value, end_color=cube[0][6].value, fill_type='solid')
    cube[4][0].value = cube[0][6].value
    cube[4][1].fill = pyxl.styles.PatternFill(start_color=cube[0][3].value, end_color=cube[0][3].value, fill_type='solid')
    cube[4][1].value = cube[0][3].value
    cube[4][2].fill = pyxl.styles.PatternFill(start_color=cube[0][0].value, end_color=cube[0][0].value, fill_type='solid')
    cube[4][2].value = cube[0][6].value
    cube[0][0].fill = pyxl.styles.PatternFill(start_color=cube[5][6].value, end_color=cube[5][6].value, fill_type='solid')
    cube[0][0].value = cube[5][6].value
    cube[0][3].fill = pyxl.styles.PatternFill(start_color=cube[5][7].value, end_color=cube[5][7].value, fill_type='solid')
    cube[0][3].value = cube[5][7].value
    cube[0][6].fill = pyxl.styles.PatternFill(start_color=cube[5][8].value, end_color=cube[5][8].value, fill_type='solid')
    cube[0][6].value = cube[5][8].value
    cube[5][6].fill = pyxl.styles.PatternFill(start_color=cube[2][8].value, end_color=cube[2][8].value, fill_type='solid')
    cube[5][6].value = cube[2][8].value
    cube[5][7].fill = pyxl.styles.PatternFill(start_color=cube[2][5].value, end_color=cube[2][5].value, fill_type='solid')
    cube[5][7].value = cube[2][5].value
    cube[5][8].fill = pyxl.styles.PatternFill(start_color=cube[2][2].value, end_color=cube[2][2].value, fill_type='solid')
    cube[5][8].value = cube[2][2].value
    cube[2][2].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[2][2].value = temp1
    cube[2][5].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[2][5].value = temp2
    cube[2][8].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[2][8].value = temp3

    #back face correction
    temp = cube[3][0].value
    cube[3][0].fill = pyxl.styles.PatternFill(start_color=cube[3][2].value, end_color=cube[3][2].value, fill_type='solid')
    cube[3][0].value = cube[3][2].value
    cube[3][2].fill = pyxl.styles.PatternFill(start_color=cube[3][8].value, end_color=cube[3][8].value, fill_type='solid')
    cube[3][2].value = cube[3][8].value
    cube[3][8].fill = pyxl.styles.PatternFill(start_color=cube[3][6].value, end_color=cube[3][6].value, fill_type='solid')
    cube[3][8].value = cube[3][6].value
    cube[3][6].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[3][6].value = temp
    temp = cube[3][1].value
    cube[3][1].fill = pyxl.styles.PatternFill(start_color=cube[3][5].value, end_color=cube[3][5].value, fill_type='solid')
    cube[3][1].value = cube[3][5].value
    cube[3][5].fill = pyxl.styles.PatternFill(start_color=cube[3][7].value, end_color=cube[3][7].value, fill_type='solid')
    cube[3][5].value = cube[3][7].value
    cube[3][7].fill = pyxl.styles.PatternFill(start_color=cube[3][3].value, end_color=cube[3][3].value, fill_type='solid')
    cube[3][7].value = cube[3][3].value
    cube[3][3].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[3][3].value = temp
    
def back_counter(cube):
    temp1 = cube[4][0].value
    temp2 = cube[4][1].value
    temp3 = cube[4][2].value
    cube[4][0].fill = pyxl.styles.PatternFill(start_color=cube[2][2].value, end_color=cube[2][2].value, fill_type='solid')
    cube[4][0].value = cube[2][2].value    
    cube[4][1].fill = pyxl.styles.PatternFill(start_color=cube[2][5].value, end_color=cube[2][5].value, fill_type='solid')
    cube[4][1].value = cube[2][5].value 
    cube[4][2].fill = pyxl.styles.PatternFill(start_color=cube[2][8].value, end_color=cube[2][8].value, fill_type='solid')
    cube[4][2].value = cube[2][8].value 
    cube[2][2].fill = pyxl.styles.PatternFill(start_color=cube[5][8].value, end_color=cube[5][8].value, fill_type='solid')
    cube[2][2].value = cube[5][8].value 
    cube[2][5].fill = pyxl.styles.PatternFill(start_color=cube[5][7].value, end_color=cube[5][7].value, fill_type='solid')
    cube[2][5].value = cube[5][7].value 
    cube[2][8].fill = pyxl.styles.PatternFill(start_color=cube[5][6].value, end_color=cube[5][6].value, fill_type='solid')
    cube[2][8].value = cube[5][6].value 
    cube[5][8].fill = pyxl.styles.PatternFill(start_color=cube[0][6].value, end_color=cube[0][6].value, fill_type='solid')
    cube[5][8].value = cube[0][6].value 
    cube[5][7].fill = pyxl.styles.PatternFill(start_color=cube[0][3].value, end_color=cube[0][3].value, fill_type='solid')
    cube[5][7].value = cube[0][3].value 
    cube[5][6].fill = pyxl.styles.PatternFill(start_color=cube[0][0].value, end_color=cube[0][0].value, fill_type='solid')
    cube[5][6].value = cube[0][0].value 
    cube[0][6].fill = pyxl.styles.PatternFill(start_color=temp1, end_color=temp1, fill_type='solid')
    cube[0][6].value = temp1
    cube[0][3].fill = pyxl.styles.PatternFill(start_color=temp2, end_color=temp2, fill_type='solid')
    cube[0][3].value = temp2
    cube[0][0].fill = pyxl.styles.PatternFill(start_color=temp3, end_color=temp3, fill_type='solid')
    cube[0][0].value = temp3

    #back face correction
    temp = cube[3][0].value
    cube[3][0].fill = pyxl.styles.PatternFill(start_color=cube[3][6].value, end_color=cube[3][6].value, fill_type='solid')
    cube[3][0].value = cube[3][6].value 
    cube[3][6].fill = pyxl.styles.PatternFill(start_color=cube[3][8].value, end_color=cube[3][8].value, fill_type='solid')
    cube[3][6].value = cube[3][8].value 
    cube[3][8].fill = pyxl.styles.PatternFill(start_color=cube[3][2].value, end_color=cube[3][2].value, fill_type='solid')
    cube[3][8].value = cube[3][2].value 
    cube[3][2].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[3][2].value = temp
    temp = cube[3][1].value
    cube[3][1].fill = pyxl.styles.PatternFill(start_color=cube[3][3].value, end_color=cube[3][3].value, fill_type='solid')
    cube[3][1].value = cube[3][3].value 
    cube[3][3].fill = pyxl.styles.PatternFill(start_color=cube[3][7].value, end_color=cube[3][7].value, fill_type='solid')
    cube[3][3].value = cube[3][7].value 
    cube[3][7].fill = pyxl.styles.PatternFill(start_color=cube[3][5].value, end_color=cube[3][5].value, fill_type='solid')
    cube[3][7].value = cube[3][5].value
    cube[3][5].fill = pyxl.styles.PatternFill(start_color=temp, end_color=temp, fill_type='solid')
    cube[3][5].value = temp


#functions below are compound moves to turn entire faces at once: up, down, left, right, clockwise, countercw

def turn_up(cube):
    left_up(cube)
    right_up(cube)
    center_up(cube)

def turn_down(cube):
    left_down(cube)
    right_down(cube)
    center_down(cube)

def turn_right(cube):
    top_right(cube)
    middle_right(cube)
    bottom_right(cube)

def turn_left(cube):
    top_left(cube)
    middle_left(cube)
    bottom_left(cube)

def turn_clockwise(cube):
    front_clockwise(cube)
    back_clockwise(cube)
    core_clockwise(cube)

def turn_counter(cube):
    front_counter(cube)
    back_counter(cube)
    core_counter(cube)        